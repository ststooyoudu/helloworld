import openpyxl
from testReadFile import DDL

class TestExcel:
    def __init__(self,xls_name,ddl_name):
        self.wb = openpyxl.load_workbook(xls_name)
        #sheet = wb.sheetnames[0]
        self.sheet = self.wb.active
        self.DDL1=DDL(ddl_name)
        try:
            for column in range(self.sheet.max_column+1):
                if column == 1:
                    self.sheet.cell(row=self.sheet.max_row+1,column=column).value=self.DDL1.Author_2
                if column == 2:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.Author_2
                if column == 3:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.Remark
                if column == 5:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= 'ddl'
                if column == 7:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.DDLAllContent
                if column == 8:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.Name
                if column == 11:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= '江苏bes'
                if column == 12:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.db_name
                if column == 13:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.db_username
                if column == 14:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= '执行一次'
                if column == 15:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= '一般'
        except:
            print(str(ddl_name)+'执行失败！已忽略！！！')
            pass
        self.wb.save('a.xlsx')
