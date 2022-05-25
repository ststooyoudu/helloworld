import openpyxl
from testReadFile import DDL
import os
from pathlib import Path

class TestExcel_NewDDL:
    def __init__(self,xls_name,ddl_name):
        self.wb = openpyxl.load_workbook(xls_name)
        #sheet = wb.sheetnames[0]
        self.sheet = self.wb.active
        self.DDL1=DDL(ddl_name)
        if self.DDL1.Create:
            for column in range(self.sheet.max_column+1):
                if column == 1:
                    self.sheet.cell(row=self.sheet.max_row+1,column=column).value='DDL'
                if column == 2:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= '618'
                if column == 3:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= 'cc'
                if column == 4:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= '永久表'
                if column == 5:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.Remark
                if column == 12:
                    self.sheet.cell(row=self.sheet.max_row,column=column).value= self.DDL1.DDLContent
            self.wb.save('/workspace/helloworld/testpython/b.xlsx')
#newexcel=TestExcel_NewDDL('/workspace/helloworld/testpython/b.xlsx','/workspace/helloworld/testpython/00483-BESCHARGE-(BESCHARGE.OM.AR_CHARGE_TASK.REQ_2022_0324002_wwx1143545.20220415192013.DDL)-Jiangsu_bes-l00293600.sql')


